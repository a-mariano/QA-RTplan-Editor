import os
import tempfile
import subprocess
import pandas as pd
from openpyxl import load_workbook
from PyQt5.QtWidgets import QMessageBox

def export_to_excel(ds):
    beams = getattr(ds, "RTBeamSequence", None) or getattr(ds, "BeamSequence", None)
    if not beams:
        QMessageBox.warning(None, "Atenção", "Nenhum feixe disponível.")
        return
    tmp_dir = tempfile.gettempdir()
    path = os.path.join(tmp_dir, "CPs_todos_beams.xlsx")
    try:
        writer = pd.ExcelWriter(path, engine="openpyxl")
        for beam_idx, beam in enumerate(beams):
            sheet_name = f"Beam_{beam_idx}"
            cps = beam.ControlPointSequence
            max_leaf_len = 0
            for cp in cps:
                bl_seq = getattr(cp, "BeamLimitingDevicePositionSequence", None)
                if bl_seq:
                    for item in bl_seq:
                        dtype = getattr(item, "RTBeamLimitingDeviceType", "").upper()
                        if "MLC" in dtype and hasattr(item, "LeafJawPositions"):
                            leaf_len = len(item.LeafJawPositions)
                            if leaf_len > max_leaf_len:
                                max_leaf_len = leaf_len
                            break
            if max_leaf_len % 2 != 0:
                max_leaf_len += 1
            half_leaf = max_leaf_len // 2
            row_labels = ["GantryAngle", "BeamLimitingDeviceAngle", "PatientSupportAngle", "CumulativeMetersetWeight"]
            for i in range(1, half_leaf+1):
                row_labels.append(f"Leaf_Left_{i}")
            for i in range(1, half_leaf+1):
                row_labels.append(f"Leaf_Right_{i}")
            data = {}
            for idx_cp, cp in enumerate(cps):
                col = f"CP{idx_cp}"
                vals = [getattr(cp, "GantryAngle", ""), getattr(cp, "BeamLimitingDeviceAngle", ""), getattr(cp, "PatientSupportAngle", ""), getattr(cp, "CumulativeMetersetWeight", "")]
                leaf_list = []
                bl_seq = getattr(cp, "BeamLimitingDevicePositionSequence", None)
                if bl_seq:
                    for item in bl_seq:
                        dtype = getattr(item, "RTBeamLimitingDeviceType", "").upper()
                        if "MLC" in dtype and hasattr(item, "LeafJawPositions"):
                            leaf_list = [float(x) for x in item.LeafJawPositions]
                            break
                left_vals = leaf_list[:half_leaf] if len(leaf_list)>=half_leaf else leaf_list+[ "" for _ in range(half_leaf-len(leaf_list))]
                right_vals = leaf_list[half_leaf:half_leaf*2] if len(leaf_list)>=half_leaf*2 else leaf_list[half_leaf:]+["" for _ in range(half_leaf-len(leaf_list[half_leaf:]))]
                if len(left_vals)<half_leaf: left_vals += ["" for _ in range(half_leaf-len(left_vals))]
                if len(right_vals)<half_leaf: right_vals += ["" for _ in range(half_leaf-len(right_vals))]
                vals.extend(left_vals); vals.extend(right_vals)
                data[col]=vals
            df = pd.DataFrame(data, index=row_labels)
            df.to_excel(writer, sheet_name=sheet_name)
        writer.close()
    except PermissionError:
        QMessageBox.warning(None, "Erro", "Não foi possível salvar Excel.")
        return
    except Exception as e:
        QMessageBox.critical(None, "Erro", f"Ocorreu erro ao gerar Excel:\n{e}")
        return
    try:
        wb = load_workbook(path)
        for sheet in wb.sheetnames:
            ws = wb[sheet]; ws.column_dimensions['A'].width=40
        wb.save(path)
    except:
        pass
    try:
        if os.name == 'nt': os.startfile(path)
        elif sys.platform=='darwin': subprocess.call(["open", path])
        else: subprocess.call(["xdg-open", path])
    except:
        QMessageBox.information(None,"Excel",f"Salvo em:\n{path}")
    QMessageBox.information(None,"Exportado",f"Excel salvo em:\n{path}")

def import_from_excel(ds, beam_idx):
    tmp_dir = tempfile.gettempdir()
    path = os.path.join(tmp_dir, "CPs_todos_beams.xlsx")
    if not os.path.exists(path):
        QMessageBox.warning(None, "Erro", "Exporte para Excel primeiro.")
        return
    sheet = f"Beam_{beam_idx}"
    try:
        df = pd.read_excel(path, sheet_name=sheet, index_col=0, engine="openpyxl")
    except Exception as e:
        QMessageBox.critical(None, "Erro", f"Não foi possível ler Excel:\n{e}"); return
    beams = getattr(ds, "RTBeamSequence", None) or getattr(ds, "BeamSequence", None)
    beam = beams[beam_idx]; cps = beam.ControlPointSequence
    existing_len = len(cps)
    cp_cols = [c for c in df.columns if c.startswith("CP")]
    total_cols = len(cp_cols)
    if total_cols > existing_len:
        last = cps[-1]
        import copy as _copy
        for i in range(existing_len,total_cols):
            cps.append(_copy.deepcopy(last))
    cps = beam.ControlPointSequence
    leaf_left_rows = [r for r in df.index if r.startswith("Leaf_Left_")]
    half_leaf = len(leaf_left_rows)
    for idx_cp, col in enumerate(cp_cols):
        cp = cps[idx_cp]
        try:
            ga = df.at["GantryAngle", col]; 
            if pd.notna(ga): cp.GantryAngle=float(ga)
            ba = df.at["BeamLimitingDeviceAngle", col]
            if pd.notna(ba): cp.BeamLimitingDeviceAngle=float(ba)
            ta = df.at["PatientSupportAngle", col]
            if pd.notna(ta): cp.PatientSupportAngle=float(ta)
            cw = df.at["CumulativeMetersetWeight", col]
            if pd.notna(cw): cp.CumulativeMetersetWeight=float(cw)
        except Exception as e:
            QMessageBox.warning(None, "Erro", f"Conversão Ângulos/Fração CP{idx_cp}:\n{e}"); return
        leaf_list=[]
        for i in range(1, half_leaf+1):
            try:
                lv = df.at[f"Leaf_Left_{i}", col]
                leaf_list.append(float(lv) if pd.notna(lv) else 0.0)
            except Exception as e:
                QMessageBox.warning(None,"Erro Leaf Left",f"CP{idx_cp} Leaf_Left_{i}:\n{e}"); return
        for i in range(1, half_leaf+1):
            try:
                lv = df.at[f"Leaf_Right_{i}", col]
                leaf_list.append(float(lv) if pd.notna(lv) else 0.0)
            except Exception as e:
                QMessageBox.warning(None,"Erro Leaf Right",f"CP{idx_cp} Leaf_Right_{i}:\n{e}"); return
        bl_seq = getattr(cp, "BeamLimitingDevicePositionSequence", None)
        mlc_item=None
        if bl_seq:
            for item in bl_seq:
                if "MLC" in getattr(item, "RTBeamLimitingDeviceType", "").upper() and hasattr(item, "LeafJawPositions"):
                    mlc_item = item; break
        if mlc_item:
            try: mlc_item.LeafJawPositions = leaf_list
            except Exception as e:
                QMessageBox.warning(None,"Erro MLC",f"Não foi possível atribuir LeafJawPositions CP{idx_cp}:\n{e}"); return
    QMessageBox.information(None, "Importado", "CPs atualizados em memória. Use 'Salvar Como...' para gravar.")
